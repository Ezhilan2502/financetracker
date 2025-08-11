from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth.decorators import login_required
from . models import Transaction,Goal,EmailOTP
from . forms import TransactionForm,GoalForm,RegisterForm,ForgetPasswordForm
from django.db.models import Sum,Q
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate,login as auth_login,logout as auth_logout
from django.contrib import messages
from django.contrib.auth.models import User
import json,random,calendar
from decimal import Decimal
from django.utils.timezone import now
from django.utils.dateparse import parse_date
from django.core.mail import send_mail
from django.urls import reverse
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models.functions import ExtractYear,ExtractMonth
from openpyxl.styles import Font

# Create your views here.

def base(request):
    return render(request,'base.html')

@login_required
def dashboard(request):

    month=request.GET.get('month')
    year=request.GET.get('year')
    if month and year:
        transactions = Transaction.objects.filter(date__month=int(month), date__year=int(year), user=request.user)
    else:
        transactions = Transaction.objects.filter(user=request.user)
        
    total_income=transactions.filter(transaction_type='IN').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense=transactions.filter(transaction_type='EX').aggregate(Sum('amount'))['amount__sum'] or 0
    savings = transactions.filter(transaction_type='SA').aggregate(Sum('amount'))['amount__sum'] or 0

    balance=total_income-total_expense-savings

    

    total_expense_by_category=(
        transactions
        .values('category')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    
    pie_chart_labels = []
    pie_chart_data = []
    color_palette = [
    "#FF6384",  # Red
    "#36A2EB",  # Blue
    "#FFCE56",  # Yellow
    "#4BC0C0",  # Teal
    "#9966FF",  # Purple
    "#FF9F40",  # Orange
    "#E7E9ED",  # Light Gray
    "#8E44AD",  # Dark Purple
    "#3498DB",  # Sky Blue
    "#1ABC9C",  # Mint
    "#2ECC71",  # Green
    "#F1C40F",  # Gold
    "#E67E22",  # Pumpkin
    "#E74C3C",  # Red
    "#95A5A6",  # Gray
    "#34495E",  # Navy
    "#D35400",  # Dark Orange
    "#7F8C8D",  # Dark Gray
    "#27AE60",  # Emerald
    "#2980B9",  # Strong Blue
    "#F39C12",  # Bright Yellow
    "#C0392B",  # Deep Red
    "#BDC3C7",  # Silver
    "#9B59B6",  # Light Purple
    "#16A085",  # Deep Teal
    "#22313F",  # Steel Blue
    "#F62459",  # Pink
    "#663399",  # Rebecca Purple
    "#FF6F61",  # Coral
    "#6B5B95",  # Plum
    ]
    for item in total_expense_by_category:
        pie_chart_labels.append(item['category'])
        pie_chart_data.append(float(item['total']))
    expense_chart_data = {
        'labels': pie_chart_labels,
        'data': pie_chart_data,
        'backgroundColor': [color_palette[i % len(color_palette)] for i in range(len(pie_chart_labels))],
    }
    expense_chart_data_json = json.dumps(expense_chart_data)

    total_goal_target=Goal.objects.filter(user=request.user).aggregate(Sum('target_amount'))['target_amount__sum'] or 0
    total_goal_target=float(total_goal_target)
    
    
    goals=Goal.objects.filter(user=request.user)
    contribution_data=[]
    for goal in goals:
        total_contrib=transactions.filter(
            transaction_type='SA',
            category='Goal Contribution',
            goal=goal
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        contribution_data.append({
            'name':goal.name,
            'saved':float(total_contrib)
        })
    total_goal_saved = sum(item['saved'] for item in contribution_data)
    progress_percentage = 0
    if total_goal_target > 0:
        progress_percentage=round((total_goal_saved/total_goal_target)*100,0)

    recent_transactions=transactions.order_by('-date')[:5]


    contribution_chart_data={
        'labels':[item['name'] for item in contribution_data],
        'data':[item['saved'] for item in contribution_data],
        'backgroundColor': [color_palette[i % len(color_palette)] for i in range(len(contribution_data))],
    }
    contribution_chart_data_json = json.dumps(contribution_chart_data)

    context={
        'total_income':float(total_income),
        'total_expense':float(total_expense),
        'savings':float(savings),
        'balance':balance,
        'selected_month': month,
        'selected_year': year,
        'expense_chart_data_json': expense_chart_data_json,
        'total_goal_target': float(total_goal_target),
        'contribution_chart_data_json': contribution_chart_data_json,
        'total_goal_saved': float(total_goal_saved),
        'progress_percentage': int(progress_percentage),
        'recent_transactions': recent_transactions,
        'expense_chart_data_json': json.dumps(expense_chart_data),
        'contribution_chart_data_json': contribution_chart_data_json,
    }
    
    return render(request,'dashboard.html',context)

@login_required
def goal_list(request):
    user_goals=Goal.objects.filter(user=request.user).order_by('target_date')
    goals_chart_data=[]
    goal_saved_color_palette=[
       "#FF6384",  # Red
    "#36A2EB",  # Blue
    "#FFCE56",  # Yellow
    "#4BC0C0",  # Teal
    "#9966FF",  # Purple
    "#FF9F40",  # Orange
    "#E7E9ED",  # Light Gray
    "#8E44AD",  # Dark Purple
    "#3498DB",  # Sky Blue
    "#1ABC9C",  # Mint
    "#2ECC71",  # Green
    "#F1C40F",  # Gold
    "#E67E22",  # Pumpkin
    "#E74C3C",  # Red
    "#95A5A6",  # Gray
    "#34495E",  # Navy
    "#D35400",  # Dark Orange
    "#7F8C8D",  # Dark Gray
    "#27AE60",  # Emerald
    "#2980B9",  # Strong Blue
    "#F39C12",  # Bright Yellow
    "#C0392B",  # Deep Red
    "#BDC3C7",  # Silver
    "#9B59B6",  # Light Purple
    "#16A085",  # Deep Teal
    "#22313F",  # Steel Blue
    "#F62459",  # Pink
    "#663399",  # Rebecca Purple
    "#FF6F61",  # Coral
    "#6B5B95",  # Plum

    ]
    remaining_color='#40E0D0'
    for i,goal in enumerate(user_goals):
        saved_amount=float(goal.current_saved_amount)
        remaining_amount=float(goal.amount_remaining)
        target_amount=float(goal.target_amount)
        saved_slice_color=goal_saved_color_palette[i%len(goal_saved_color_palette)]
        goal_data_for_chart={
            'id':goal.id,
            'name':goal.name,
            'target_amount':target_amount,
            'current_saved_amount':saved_amount,
            'target_date':goal.target_date.strftime('%Y-%m-%d') if goal.target_date else 'N/A',
            'progress_percentage': float(round(goal.percentage_complete, 2)),
            'chart_data': {
                'labels': ['Saved', 'Remaining'],
                'datasets': [{
                    'data': [saved_amount, remaining_amount],
                    'backgroundColor': [saved_slice_color, remaining_color],
                    'hoverOffset': 4
                }]
            }

        }
        goals_chart_data.append(goal_data_for_chart)

    goals_chart_data_json=json.dumps(goals_chart_data)
    context={
        'goals':user_goals,
        'goals_chart_data_json':goals_chart_data_json,
    }
    return render(request,'goal_list.html',context)

def message_redirect(request):
    redirect_url=request.GET.get('next','dashboard')
    return render(request,'message_redirect.html',{'redirect_url':redirect_url})

@login_required
def add_goal(request):
    if request.method == 'POST':
        form = GoalForm(request.POST)
        if form.is_valid():
            goal = form.save(commit=False)
            goal.user = request.user
            goal.save()
            messages.success(request, "Goal created successfully!")
            
            return redirect(reverse('message_redirect') + '?next=' + reverse('goal_list'))
    else:
        form = GoalForm()
    return render(request, 'add_goal.html', {'form': form})

@login_required
def edit_goal(request, pk):
    goal = get_object_or_404(Goal, pk=pk, user=request.user)
    if request.method == 'POST':
        form = GoalForm(request.POST, instance=goal)
        if form.is_valid():
            form.save()
            messages.success(request, "Goal updated successfully!")
            return redirect(reverse('message_redirect') + '?next=' + reverse('goal_list'))
    else:
        form = GoalForm(instance=goal)
    return render(request, 'edit_goal.html', {'form': form,'goal':goal})

@login_required
def delete_goal(request, pk):
    goal = get_object_or_404(Goal, pk=pk, user=request.user)
    if request.method == 'POST':
        Transaction.objects.filter(goal=goal).delete()
        goal.delete()
        messages.success(request, "Goal deleted successfully!")
        return redirect(reverse('message_redirect') + '?next=' + reverse('goal_list'))
    return render(request, 'delete_goal.html', {'goal': goal})

@login_required
def contribute_to_goal(request,pk):
    goal=get_object_or_404(Goal,pk=pk,user=request.user)
    if request.method=='POST':
        try:
            contribution_amount=Decimal(request.POST.get('amount'))
            if contribution_amount<=0:
                messages.error(request,"Contribution must be positive")
                return render(request,'contribute_to_goal.html',{'goal':goal})
            if goal.current_saved_amount+contribution_amount>goal.target_amount:
                messages.error(request,"Contribution exceeds the goal's target amount")
                return render(request,'contribute_to_goal',pk=goal.pk)
            remaining_goal=goal.target_amount-goal.current_saved_amount
            if contribution_amount>remaining_goal:
                messages.error(request,f"Contribution exceeds the goal's remaining amount (₹{remaining_goal})")
                return render(request,'contribute_to_goal.html',{'goal':goal})

            
            total_income=Transaction.objects.filter(user=request.user,transaction_type='IN').aggregate(Sum('amount'))['amount__sum'] or 0
            total_expense = Transaction.objects.filter(user=request.user, transaction_type='EX').aggregate(Sum('amount'))['amount__sum'] or 0
            total_savings = Transaction.objects.filter(user=request.user, transaction_type='SA').aggregate(Sum('amount'))['amount__sum'] or 0

            available_balance=total_income-total_expense-total_savings

            if contribution_amount>available_balance:
                messages.error(request, f"Insufficient balance. You only have ₹{available_balance} available.")
                return redirect(reverse('message_redirect') + '?next=' + reverse('contribute_to_goal', args=[goal.pk]))
               
            goal.current_saved_amount+=contribution_amount
            goal.save()

            Transaction.objects.create(
                user=request.user,
                title=f"Contribution to {goal.name}",
                amount=contribution_amount,
                transaction_type='SA',
                category='Goal Contribution',
                date=now().date(),
                goal=goal
            )
            messages.success(request,f"Contributed ₹{contribution_amount} to '{goal.name}'. ")
            return redirect(reverse('message_redirect') + '?next=' + reverse('goal_list'))

                

            
        except Exception as e:
            messages.error(request,f"Invalid amount:{contribution_amount}")
        return redirect(reverse('message_redirect') + '?next=' + reverse('goal_list'))
    return render(request,'contribute_to_goal.html',{'goal':goal})


def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            phone=form.cleaned_data['phone']
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already registered.')
                return redirect(reverse('message_redirect') + '?next=' + reverse('register'))
            elif User.objects.filter(phone=phone).exists():
                messages.error(request,'Phone no already registered')
                return redirect(reverse('message_redirect') + '?next=' + reverse('register'))
            
            otp = str(random.randint(100000, 999999))
            EmailOTP.objects.create(email=email, otp=otp)

            send_mail(
                'Your OTP for Registration',
                f'Your OTP is {otp}',
                'ezhilanarunachalam25@gmail.com',
                [email],
                fail_silently=False,
            )

            request.session['reg_data'] = {
                'name': form.cleaned_data['name'],
                'phone': form.cleaned_data['phone'],
                'email': email,
                'password': form.cleaned_data['password'],
            }
            messages.success(request, 'OTP sent to your email. Please verify.')
            return redirect(reverse('message_redirect') + '?next=' + reverse('verify_otp'))

    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form})


def verify_otp(request):
    if request.method == 'POST':
        entered_otp = request.POST.get('otp')
        reg_data = request.session.get('reg_data')

        if not reg_data:
            messages.error(request, "Session expired. Please register again.")
            return redirect(reverse('message_redirect') + '?next=' + reverse('register'))

        try:
            email_otp = EmailOTP.objects.filter(email=reg_data['email']).latest('created_at')
        except EmailOTP.DoesNotExist:
            messages.error(request, "No OTP found. Try again.")
            return redirect(reverse('message_redirect') + '?next=' + reverse('register'))

        if email_otp.otp == entered_otp and not email_otp.is_expired():
            
            user = User.objects.create_user(
                username=reg_data['email'],
                email=reg_data['email'],
                password=reg_data['password'],
                first_name=reg_data['name'],
            )
           
            messages.success(request, 'Registration successful! Please log in.')
            del request.session['reg_data']
            return redirect(reverse('message_redirect') + '?next=' + reverse('login'))
        else:
            messages.error(request, 'Invalid or expired OTP.')

    return render(request, 'verify_otp.html')

def login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email') 
        password = request.POST.get('password')
        

        try:
            user_obj = User.objects.get(email=email)
            username = user_obj.username
            user = authenticate(request, username=username, password=password)
        except User.DoesNotExist:
            user = None

        if user is not None:
            auth_login(request, user)
            messages.success(request,'Login Successfully')
            return redirect(reverse('message_redirect') + '?next=' + reverse('dashboard'))
        else:
            messages.error(request, "Invalid Email or Password")
            return redirect(reverse('message_redirect') + '?next=' + reverse('login'))

    return render(request, 'login.html')
def generate_otp():
    return str(random.randint(100000,999999))

def forget_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, "No account found with this email.")
            return redirect('forget_password')
        
        otp = generate_otp()
        EmailOTP.objects.create(email=email, otp=otp)

        send_mail(
            'Your OTP for Password Reset',
            f'Your OTP is: {otp}',
            'youremail@gmail.com',
            [email],
            fail_silently=False
        )
        request.session['reset_email'] = email
        return redirect(reverse('message_redirect') + '?next=' + reverse('verify_otp1'))

    return render(request, 'forget_password.html')

def verify_otp1(request):
    if request.method == 'POST':
        email = request.session.get('reset_email')
        otp_entered = request.POST.get('otp')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect(reverse('message_redirect') + '?next=' + reverse('verify_otp1'))

        try:
            otp_obj = EmailOTP.objects.filter(email=email).latest('created_at')
        except EmailOTP.DoesNotExist:
            messages.error(request, "No OTP found.")
            return redirect(reverse('message_redirect') + '?next=' + reverse('forget_password'))

        if otp_obj.otp != otp_entered:
            messages.error(request, "Invalid OTP.")
            return redirect(reverse('message_redirect') + '?next=' + reverse('verify_otp1'))

        if otp_obj.is_expired():
            messages.error(request, "OTP expired. Please try again.")
            return redirect(reverse('message_redirect') + '?next=' + reverse('forget_password'))

        user = User.objects.get(email=email)
        user.set_password(new_password)
        user.save()
        messages.success(request, "Password reset successful. Please login.")
        return redirect(reverse('message_redirect') + '?next=' + reverse('login'))

    return render(request, 'verify_otp1.html')

def logout_view(request):
    auth_logout(request)
    messages.success(request,"Logout successfully")
    return redirect(reverse('message_redirect') + '?next=' + reverse('login'))

@login_required
def transaction_list(request):

    transactions=Transaction.objects.filter(user=request.user)
    
    start_date=request.GET.get('start_date')
    end_date=request.GET.get('end_date')
    category=request.GET.get('category')
    transaction_type=request.GET.get('transaction_type')

    if start_date:
        transactions=transactions.filter(date__gte=parse_date(start_date))

    if end_date:
        transactions=transactions.filter(date__lte=parse_date(end_date))
    if category and category!='all':
        transactions=transactions.filter(category=category)
    if transaction_type and transaction_type!='all':
        transactions=transactions.filter(transaction_type=transaction_type)

    category_sums=(
        transactions
        .exclude(category__isnull=True)
        .exclude(category__exact='')
        .values('category')
        .annotate(total=Sum('amount'))
        .order_by('category')
    )
    categories = Transaction.objects.filter(user=request.user) \
                .exclude(category__isnull=True) \
                .exclude(category__exact='') \
                .values_list('category', flat=True).distinct()

    labels=[]
    data=[]
    for entry in category_sums:
        labels.append(entry['category'])
        data.append(float(entry['total']))

    background_colors = [
       "#FF6384",  # Red
    "#36A2EB",  # Blue
    "#FFCE56",  # Yellow
    "#4BC0C0",  # Teal
    "#9966FF",  # Purple
    "#FF9F40",  # Orange
    "#E7E9ED",  # Light Gray
    "#8E44AD",  # Dark Purple
    "#3498DB",  # Sky Blue
    "#1ABC9C",  # Mint
    "#2ECC71",  # Green
    "#F1C40F",  # Gold
    "#E67E22",  # Pumpkin
    "#E74C3C",  # Red
    "#95A5A6",  # Gray
    "#34495E",  # Navy
    "#D35400",  # Dark Orange
    "#7F8C8D",  # Dark Gray
    "#27AE60",  # Emerald
    "#2980B9",  # Strong Blue
    "#F39C12",  # Bright Yellow
    "#C0392B",  # Deep Red
    "#BDC3C7",  # Silver
    "#9B59B6",  # Light Purple
    "#16A085",  # Deep Teal
    "#22313F",  # Steel Blue
    "#F62459",  # Pink
    "#663399",  # Rebecca Purple
    "#FF6F61",  # Coral
    "#6B5B95",  # Plum
    ][:len(labels)]

    expense_chart_data_json=json.dumps({
        'labels':labels,
        'data':data,
        'backgroundColor':background_colors,
    })
    context = {
        'transactions': transactions.order_by('-date'),
        'categories': categories,       
        'transaction_types': Transaction.TRANSACTION_TYPES,
        'filters': {
            'start_date': start_date or '',
            'end_date': end_date or '',
            'category': category or 'all',
            'transaction_type': transaction_type or 'all',
        },
        'expense_chart_data_json': expense_chart_data_json,
    }
    
    return render(request,'transaction_list.html',context)

@login_required
def add_transaction(request):
    if request.method=='POST':
        form=TransactionForm(request.POST)
        if form.is_valid():
            transaction=form.save(commit=False)
            transaction.user=request.user
            transaction.save()
            messages.success(request,"Transaction created succesfully")
            return redirect(reverse('message_redirect') + '?next=' + reverse('transaction_list'))
        else:
            messages.error(request, "Please correct the errors in the form.")
            return redirect(reverse('message_redirect') + '?next=' + reverse('transaction_list'))
            
    else:
        form=TransactionForm()
    return render(request,'add_transaction.html',{'form':form})
        
@login_required
def edit_transaction(request,pk):
    transaction=get_object_or_404(Transaction,pk=pk,user=request.user)
    if request.method=='POST':
        form=TransactionForm(request.POST,instance=transaction)
        if form.is_valid():
            updated_transaction=form.save()
            if (updated_transaction.transaction_type=='SA' and 
                updated_transaction.category=='Goal Contribution' and
                updated_transaction.title.startswith("Contribution to ")):
                goal_name=updated_transaction.title.replace("Contribution to ","").strip()
                try:
                    goal=Goal.objects.get(user=request.user,name=goal_name)
                    total_contrib=Transaction.objects.filter(
                        user=request.user,
                        transaction_type='SA',
                        category='Goal Contribution',
                        goal=goal
                    ).aggregate(Sum('amount'))['amount__sum'] or 0
                    goal.current_saved_amount=total_contrib
                    goal.save()
                   
                except Goal.DoesNotExist:
                    pass

            
            messages.success(request,"Transaction Edited succesfully")
            return redirect(reverse('message_redirect') + '?next=' + reverse('transaction_list'))
    else:
        form=TransactionForm(instance=transaction)
    return render(request,'edit_transaction.html',{'form':form})

@login_required
def delete_transaction(request,pk):
    transaction=get_object_or_404(Transaction,pk=pk,user=request.user)

    is_goal_contribution=(
        transaction.transaction_type=='SA' and
        transaction.category=='Goal Contribution' and
        transaction.title.startswith("Contribution to")
    )
    goal_name=None
    if is_goal_contribution:
        goal_name=transaction.title.replace("Contribution to ","").strip()
    
    if request.method=='POST':
        transaction.delete()
        if is_goal_contribution and goal_name:
            try:
                goal=Goal.objects.get(user=request.user,name=goal_name)
                total_contrib=Transaction.objects.filter(
                    user=request.user,
                    transaction_type='SA',
                    category='Goal Contribution',
                    goal=goal

                ).aggregate(Sum('amount'))['amount__sum'] or 0
                goal.current_saved_amount=total_contrib
                goal.save()
                
            except Goal.DoesNotExist:
                pass
        messages.success(request,"Transaction deleted successfully")
        return redirect(reverse('message_redirect') + '?next=' + reverse('transaction_list'))
    return render(request,'delete_transaction.html',{'transaction':transaction}) 


def monthly_export(request):
    user=request.user
    monthly_data=(
        Transaction.objects.filter(user=user)
        .annotate(
            year=ExtractYear('date'),
            month=ExtractMonth('date')
        )
        .values('year','month')
        .annotate(
            income=Sum('amount',filter=Q(transaction_type='IN')),
            expenses=Sum('amount',filter=Q(transaction_type='EX')),
            savings=Sum('amount',filter=Q(transaction_type='SA')),
        )
        .order_by('year','month')
    )
    wb=Workbook()
    ws=wb.active
    ws.title='Monthly Summary'

    bold_font=Font(bold=True)

    headers=['SL.No','Month','Year','Income','Expenses','Savings','Balance']
    ws.append(headers)
    for col_num in range(1,len(headers)+1):
        ws.cell(row=1,column=col_num).font=bold_font
    total_income=0
    total_expenses=0
    total_savings=0
    total_balance=0

    for idx,data in enumerate(monthly_data,start=2):
        month_name=calendar.month_name[data['month']] if data['month'] else 'N/A'
        year=data['year'] or 'N/A'
        income=data['income'] or 0
        expenses=data['expenses'] or 0
        savings=data['savings'] or 0
        balance=income-expenses-savings

        ws.append([idx-1,month_name,year,income,expenses,savings,balance])

        total_income+=income
        total_balance+=balance
        total_expenses+=expenses
        total_savings+=savings
    total_row_num=ws.max_row+2

    ws.append([])
    totals=['','Total','',total_income,total_expenses,total_savings,total_balance]
    ws.append(totals)

    for col_num in range(1,len(totals)+1):
        ws.cell(row=total_row_num,column=col_num).font=bold_font
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Monthly_summary.xlsx'

    wb.save(response)
    return response
